#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
通过openpyxl模块操作excel的类,需要提前安装openpyxl
注释掉openpyxl/reader/worksheet.py中的250+行的
-        for cr in merged.mergeCell:
-            self.ws._clean_merge_range(cr)
"""

from openpyxl.chart import LineChart, PieChart, BarChart, ScatterChart, DoughnutChart, Reference, Series
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis
import openpyxl


class ExcelOpenPyXl(object):
    """
    @public: 自定义excel操作使用openpyxl模块
    """

    def __init__(self, input_excel_file=None):
        """
        @private: 自执行函数
        @param input_excel_file: 打开的excel文件,传则是修改或读取,需要填写文件全路径,不传则是新增
        """
        self.__all_sheet_list = list()
        # noinspection PyBroadException
        try:
            self.__input_excel_file = input_excel_file
            if self.__input_excel_file:
                self.__workbook_obj = openpyxl.load_workbook(self.__input_excel_file)

                self.reason_var = "读取excel成功"
            else:
                self.__workbook_obj = openpyxl.Workbook()
                self.reason_var = "创建空excel成功"
            self.get_total_sheet()
            self.result_str = "pass"
        except Exception as error_message:
            self.result_str = "fail"
            self.reason_var = f"读取excel失败,读取文件为:{input_excel_file},原因:{error_message}"
        # 当前使用的sheet对象,初始为第一个sheet
        self.__worksheet_obj = self.__workbook_obj.active
        #  边框
        self.__border_obj = Border(
            left=Side(border_style='thin', color='FF000000'), right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'), bottom=Side(border_style='thin', color='FF000000'))
        self.__fill_dict = dict()
        self.__font_dict = dict()
        self.__align_dict = dict()
        # 文字居中,horizontal水平方向,vertical垂直方向wrap_text自动换行,True/False
        self.__align_dict["all_center"] = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.__align_dict["center_top"] = Alignment(horizontal='center', vertical='top', wrap_text=True)
        self.__align_dict["center_bottom"] = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        self.__align_dict["left_center"] = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.__align_dict["right_center"] = Alignment(horizontal='right', vertical='center', wrap_text=True)
        # 当前单元格的值
        self.cell_value_str = None
        self.__close_bool = False

    def __str__(self):
        return "自定义excel操作模块"

    def __del__(self):
        self.close_excel()

    def close_excel(self):
        """
        @public: 关闭excel
        @return: 无
        """
        if not self.__close_bool:
            # noinspection PyBroadException
            try:
                self.__workbook_obj.close()
                self.reason_var = "关闭excel成功"
                self.result_str = "pass"
                self.__close_bool = True
            except Exception as error_message:
                self.reason_var = f"关闭excel失败,原因:{error_message}"
                self.result_str = "fail"
        return self

    def save_excel(self, input_save_file=None):
        """
        @public: 保存excel
        @param input_save_file: 保存的文件名,需要填写文件全路径,不填则使用默认,如果是新增文件则没有默认值
        @return: 无
        """
        if not input_save_file:
            if not self.__input_excel_file:
                self.result_str = "fail"
                self.reason_var = "文件名没有,无法保存"
            else:
                input_save_file = self.__input_excel_file
        if self.result_str == "pass":
            # noinspection PyBroadException
            try:
                self.__workbook_obj.save(input_save_file)
                self.result_str = "pass"
                self.reason_var = "添加sheet成功"
            except Exception as error_message:
                self.result_str = "fail"
                self.reason_var = f"保存excel失败,名称为:{input_save_file},原因:{error_message}"
        return self

    def get_total_sheet(self):
        """
        @public: 读取当前excel全部sheet
        @return: 无
        """
        self.__all_sheet_list = self.__workbook_obj.sheetnames
        return self.__all_sheet_list

    def get_sheet(self, input_sheet_name=None):
        """
        @public: 读取sheet对象
        @param input_sheet_name: 读取sheet名称,不填则使用当前默认
        @return: 返回sheet对象
        """
        if input_sheet_name:
            if input_sheet_name in self.__all_sheet_list:
                worksheet_obj = self.__workbook_obj[input_sheet_name]
            else:
                self.get_total_sheet()
                if input_sheet_name in self.__all_sheet_list:
                    worksheet_obj = self.__workbook_obj[input_sheet_name]
                else:
                    self.result_str = "fail"
                    self.reason_var = f"输入的sheet名:{input_sheet_name},未找到"
                    worksheet_obj = None
        else:
            worksheet_obj = self.__worksheet_obj
        return worksheet_obj

    def switch_sheet(self, input_sheet_name):
        """
        @public: 切换到指定sheet
        @param input_sheet_name: 需要切换的sheet名称
        @return: 无
        """
        self.__worksheet_obj = self.get_sheet(input_sheet_name)
        return self

    def add_sheet(self, input_sheet_name, index_num=0):
        """
        @public: 增加sheet
        @param input_sheet_name: 需要添加的sheet名称
        @param index_num: 添加sheet的位置默认最前面
        @return: 无
        """
        # noinspection PyBroadException
        try:
            # index_num 从0开始是第一个sheet
            self.__workbook_obj.create_sheet(title=input_sheet_name, index=index_num)
            self.result_str = "pass"
            self.reason_var = "添加sheet成功"
        except Exception as error_message:
            self.result_str = "fail"
            self.reason_var = f"添加sheet失败,sheet为:{input_sheet_name},原因:{error_message}"
        return self

    def del_sheet(self, input_sheet_name=None):
        """
        @public: 删除sheet
        @param input_sheet_name: 删除的sheet名称
        @return: 无
        """
        # noinspection PyBroadException
        try:
            if not input_sheet_name:
                obj_del = self.__worksheet_obj
            else:
                obj_del = self.__workbook_obj[input_sheet_name]
            self.__workbook_obj.remove(obj_del)
            self.result_str = "pass"
            self.reason_var = "删除sheet成功"
        except Exception as error_message:
            self.result_str = "fail"
            self.reason_var = f"删除sheet失败,原因:{error_message}"
        return self

    def modify_sheet(self, new_sheet_name, old_sheet_name=None):
        """
        @public: 修改sheet名称
        @param new_sheet_name: 改后的sheet名
        @param old_sheet_name: 改前的sheet名
        @return:
        """
        # noinspection PyBroadException
        try:
            if not old_sheet_name:
                modify_sheet_obj = self.__worksheet_obj
            else:
                modify_sheet_obj = self.__workbook_obj[old_sheet_name]
            modify_sheet_obj.title = new_sheet_name
            self.result_str = "pass"
            self.reason_var = "添加sheet成功"
        except Exception as error_message:
            self.result_str = "fail"
            self.reason_var = f"修改sheet名称失败,想要修改为为:{new_sheet_name},原因:{error_message}"
        return self

    def freeze_panes(self, row_num, col_num, input_sheet_name=None):
        """
        @public: 冻结单元格
        @param row_num: 冻结的行号
        @param col_num: 冻结的列号
        @param input_sheet_name: sheet名称,不填则使用当前默认
        @return: 无
        """
        panes_str = self.__get_cell_id(row_num, col_num)
        if self.result_str == "pass":
            worksheet_obj = self.get_sheet(input_sheet_name)
            if self.result_str == "pass":
                # noinspection PyBroadException
                try:
                    worksheet_obj.freeze_panes = panes_str
                    self.result_str = "pass"
                    self.reason_var = "冻结窗口成功"
                except Exception as error_message:
                    self.result_str = "fail"
                    self.reason_var = f"冻结窗口失败,原因:{error_message}"
        return self

    def set_col_dimensions(self, col_num, size_num, input_sheet_name=None):
        """
        @public: 设置列宽
        @param col_num: 列号
        @param size_num: 宽度值
        @param input_sheet_name: sheet名称,不填则使用当前默认
        @return: 无
        """
        worksheet_obj = self.get_sheet(input_sheet_name)
        if self.result_str == "pass":
            # noinspection PyBroadException
            try:
                worksheet_obj.column_dimensions[col_num].width = size_num
                self.result_str = "pass"
                self.reason_var = f"设置列{col_num}的宽度{size_num}成功"
            except Exception as error_message:
                self.result_str = "fail"
                self.reason_var = f"设置列宽失败,原因:{error_message}"
        return self

    def set_row_dimensions(self, row_num, size_num, input_sheet_name=None):
        """
        @public: 设置行高
        @param row_num: 行号
        @param size_num: 高度值
        @param input_sheet_name: sheet名称,不填则使用当前默认
        @return: 无
        """
        worksheet_obj = self.get_sheet(input_sheet_name)
        if self.result_str == "pass":
            # noinspection PyBroadException
            try:
                worksheet_obj.row_dimensions[row_num].height = size_num
                self.result_str = "pass"
                self.reason_var = f"设置行{row_num}的高度{size_num}成功"
            except Exception as error_message:
                self.result_str = "fail"
                self.reason_var = f"设置行高失败,原因:{error_message}"
        return self

    def set_filter(self, col_start_num, col_end_num, input_sheet_name=None):
        """
        @public: 设置过滤器
        @param col_start_num: 起始列号
        @param col_end_num: 结束列号
        @param input_sheet_name: sheet名称,不填则使用当前默认
        @return: 无
        """
        col_start_str = self.__change_col(col_start_num)
        if self.result_str == "pass":
            col_end_str = self.__change_col(col_end_num)
            if self.result_str == "pass":
                str_col_value = f"{col_start_str}:{col_end_str}"
                worksheet_obj = self.get_sheet(input_sheet_name)
                if self.result_str == "pass":
                    # noinspection PyBroadException
                    try:
                        worksheet_obj.auto_filter.ref = str_col_value
                        self.result_str = "pass"
                        self.reason_var = "设置filter成功"
                    except Exception as error_message:
                        self.reason_var = f"设置filter失败,原因:{error_message}"
                        self.result_str = "fail"
        return self

    def get_max_row_col(self, row_bool=True, col_bool=True, input_sheet_name=None):
        """
        @public: 读取最大行数,列数
        @param row_bool: 是否读取最大行数
        @param col_bool: 是否读取最大列数
        @param input_sheet_name: 指定sheet名称,不指定则用当前默认
        @return: 最大行数,最大列数
        """
        worksheet_obj = self.get_sheet(input_sheet_name)
        if self.result_str == "pass":
            if row_bool:
                row_max_num = worksheet_obj.max_row
            else:
                row_max_num = 0
            if col_bool:
                col_max_num = worksheet_obj.max_column
            else:
                col_max_num = 0
        else:
            # 行列都是从1开始,因此0不存在
            row_max_num = 0
            col_max_num = 0
        if row_bool and not col_bool:
            return row_max_num
        elif col_bool and not row_bool:
            return col_max_num
        else:
            return row_max_num, col_max_num

    def read_cell_value(self, row_num, col_num, input_sheet_name=None):
        """
        @public: 读取单元格内容
        @param row_num: 行号
        @param col_num: 列号
        @param input_sheet_name: 指定sheet名称,不指定则用当前默认
        @return: 读出来的单元格内容,不取用也可以读取类中变量cell_value_str
        """
        worksheet_obj = self.get_sheet(input_sheet_name)
        if self.result_str == "pass":
            # noinspection PyBroadException
            try:
                var_cell_value = worksheet_obj.cell(row=row_num, column=col_num).value
                if var_cell_value or var_cell_value == 0:
                    self.cell_value_str = str(var_cell_value)
                    self.reason_var = "读取单元格成功"
                else:
                    self.cell_value_str = None
                    self.reason_var = "单元格内容为空"
                self.result_str = "pass"
            except Exception as error_message:
                self.result_str = "fail"
                self.reason_var = f"读取单元格失败,单元格为:{row_num}行,{col_num}列,原因:{error_message}"
                self.cell_value_str = None
        return self.cell_value_str

    def write_cell(self, row_num, col_num, set_value_var, input_sheet_name=None,
                   fill_color_str=None, font_color_str="000000", bold_bool=False, size_num=12,
                   align_str="all_center", border_bool=False):
        """
        @public: 写入单个单元格
        @param row_num: 行号
        @param col_num: 列号
        @param set_value_var: 写入内容
        @param input_sheet_name: 指定sheet名称,不指定则用当前默认
        @param fill_color_str: 背景色,不传则不设置
        @param font_color_str: 字体颜色,不设置则为黑色
        @param bold_bool: 是否加粗,默认不加粗
        @param size_num: 字号,默认12
        @param align_str: 对齐方式,默认居中
        @param border_bool: 是否有边框,默认无边框
        @return: 无
        """
        cell_value_str = self.__get_cell_id(row_num, col_num)
        self.__set_cell_value(set_value_var, cell_value_str, input_sheet_name=input_sheet_name)
        if self.result_str == "pass":
            self.__set_cell_property(cell_value_str, input_sheet_name, fill_color_str,
                                     font_color_str, bold_bool, size_num, align_str, border_bool)
        return self

    def merge_cells(self, row_start_num, col_start_num, row_end_num, col_end_num, set_value_var,
                    input_sheet_name=None, fill_color_str=None, font_color_str="00000000", bold_bool=False,
                    size_num=12, align_str="all_center", border_bool=False):
        """
        @public: 合并写入单元格
        @param row_start_num: 起始行号
        @param col_start_num: 起始列号
        @param row_end_num: 结束行号
        @param col_end_num: 结束列号
        @param set_value_var: 写入内容
        @param input_sheet_name: 指定sheet名称,不指定则用当前默认
        @param fill_color_str: 背景色,不传则不设置
        @param font_color_str: 字体颜色,不设置则为黑色
        @param bold_bool: 是否加粗,默认不加粗
        @param size_num: 字号,默认12
        @param align_str: 对齐方式,默认居中
        @param border_bool: 是否有边框,默认无边框
        @return: 无
        """
        cell_start_str = self.__get_cell_id(row_start_num, col_start_num)
        cell_end_str = self.__get_cell_id(row_end_num, col_end_num)
        self.__set_cell_value(set_value_var, cell_start_str, cell_end_str, input_sheet_name)
        if self.result_str == "pass":
            self.__set_cell_property(cell_start_str, input_sheet_name, fill_color_str,
                                     font_color_str, bold_bool, size_num, align_str, border_bool)
        return self

    def write_list_cells(self, write_list, input_sheet_name=None):
        """
        @public: 根据list插入sheet页
        @param write_list: 数据的list,格式为:[[x,y,z], [1,2,3], [1,2,3]]
        @param input_sheet_name: 指定sheet名称,不指定则用当前默认
        @return: 无
        """
        worksheet_obj = self.get_sheet(input_sheet_name)
        for write_value in write_list:
            worksheet_obj.append(write_value)

    def add_img_to_sheet(self, img_file, row_num, col_num, img_width_num, img_height_num, input_sheet_name=None):
        """
        @public: 插入图片
        @param img_file: 图片文件全路径
        @param row_num: 插入单元格的行号
        @param col_num: 插如单元格的列号
        @param img_width_num: 图片宽
        @param img_height_num: 图片高
        @param input_sheet_name: 指定sheet名称,不指定则用当前默认
        @return: 无
        """
        # 自带图表很多不支持,没支持的用 matplotlib生成后,再插如excel
        worksheet_obj = self.get_sheet(input_sheet_name)
        if self.result_str == "pass":
            cell_value_str = self.__get_cell_id(row_num, col_num)
            img_obj = Image(img_file)
            img_obj.width = img_width_num
            img_obj.height = img_height_num
            worksheet_obj.add_image(img_obj, cell_value_str)
        return self

    def __change_col(self, input_col_var):
        """
        @private: excel的列数字,字母互转
        @param input_col_var: 列的ID值,传进来是数字,则转字母,传进来字母则转数字
        @return: 转换后的值
        """
        value_type = type(input_col_var)
        if value_type is int or input_col_var.isdigit():
            mode_bool = True
            if value_type is not int:
                input_col_var = int(input_col_var)
        else:
            mode_bool = False
        # noinspection PyBroadException
        try:
            if mode_bool:
                return_value = get_column_letter(input_col_var)
            else:
                # 这个暂时没用
                return_value = column_index_from_string(input_col_var)
            self.result_str = "pass"
            self.reason_var = "转换成功"
        except Exception as error_message:
            return_value = None
            self.result_str = "fail"
            self.reason_var = f"{input_col_var},转换失败,原因:{error_message}"
        return return_value

    def __get_cell_id(self, row_value, col_value):
        """
        @private: 单元格ID生成
        @param row_value: 行的id,最好是数字
        @param col_value: 列的id,最好是数字
        @return: 生成好可以直接只用的单元格id值,如B2
        """
        col_str = self.__change_col(col_value)
        if self.result_str == "pass":
            cell_value_str = f"{col_str}{row_value}"
        else:
            cell_value_str = None
        return cell_value_str

    def __get_fill(self, color_str="00FFFFFF"):
        """
        @private: 读取并设置单元格背景色
        @param color_str: 背景色值
        @return: 背景色对应的实例
        """
        if color_str not in self.__fill_dict.keys():
            self.__fill_dict[color_str] = PatternFill(fill_type='solid', fgColor=color_str)
        return self.__fill_dict[color_str]

    def __get_font(self, color_str="000000", bold_bool=False, size_num=12):
        """
        @private: 读取并设置单元格文字属性
        @param color_str: 字体颜色
        @param bold_bool: 是否加粗
        @param size_num: 字号值
        @return: 单元格文字对应的实例
        """
        key_str = f"{color_str}_{size_num}_{bold_bool}"
        if key_str not in self.__font_dict.keys():
            self.__font_dict[key_str] = Font(size=size_num, color=color_str, bold=bold_bool)
        return self.__font_dict[key_str]

    def __get_align(self, align_str="all_center"):
        """
        @private: 读取并设置单元格对齐方式
        @param align_str: 对齐方式
        @return: 单元格文字对应的实例
        """
        if align_str not in self.__align_dict.keys():
            return_obj = self.__align_dict["all_center"]
        else:
            return_obj = self.__align_dict[align_str]
        return return_obj

    def __set_cell_property(self, input_cell_str, input_sheet_name=None, fill_color_str=None, font_color_str="000000",
                            bold_bool=False, size_num=12, align_str="all_center", border_bool=False):
        """
        @private: 设置单元格属性
        @param input_cell_str: 单元格值
        @param input_sheet_name: 指定sheet名称,不指定则用当前默认
        @param fill_color_str: 背景色,不传则不设置
        @param font_color_str: 字体颜色,不设置则为黑色
        @param bold_bool: 是否加粗,默认不加粗
        @param size_num: 字号,默认12
        @param align_str: 对齐方式,默认居中
        @param border_bool: 是否有边框,默认无边框
        @return: 无
        """
        if self.result_str == "pass":
            worksheet_obj = self.get_sheet(input_sheet_name)
            if self.result_str == "pass":
                worksheet_obj[input_cell_str].alignment = self.__get_align(align_str)
                worksheet_obj[input_cell_str].font = self.__get_font(font_color_str, bold_bool, size_num)
                if fill_color_str:
                    worksheet_obj[input_cell_str].fill = self.__get_fill(fill_color_str)
                if border_bool:
                    worksheet_obj[input_cell_str].border = self.__border_obj
        return self

    def __set_cell_value(self, set_value_var, start_cell_str, end_cell_str=None, input_sheet_name=None):
        """
        @private: 设置单元格文字
        @param set_value_var: 写入内容
        @param start_cell_str: 起始单元格
        @param end_cell_str: 结束单元格
        @param input_sheet_name: 指定sheet名称,不指定则用当前默认
        @return: 无
        """
        worksheet_obj = self.get_sheet(input_sheet_name)
        if self.result_str == "pass":
            # noinspection PyBroadException
            try:
                if end_cell_str:
                    worksheet_obj.merge_cells(f"{start_cell_str}:{end_cell_str}")
                worksheet_obj[start_cell_str] = set_value_var
                self.result_str = "pass"
                self.reason_var = "写入单元格成功"
            except Exception as error_message:
                self.result_str = "fail"
                if end_cell_str:
                    value_str = f"合并单元格写入失败,单元格:{start_cell_str}到{end_cell_str}"
                else:
                    value_str = f"写入单元格失败,单元格:{start_cell_str}"
                self.reason_var = f"{value_str},写入值:{set_value_var},原因:{error_message}"
        return self

    def __make_chart(self, chart_type, row_chart_num, col_chart_num,
                     row_label_min_num, row_label_max_num, col_label_min_num,
                     row_data_min_num, row_data_max_num, col_data_min_num, col_data_max_num=None,
                     chart_title=None, chart_x_title=None, chart_y_title=None, chart_style_num=10,
                     input_sheet_name=None, color_list=None, symbol_str="circle", smooth_bool=False,
                     chart_width_num=None, chart_height_num=None):
        """
        @private: 生成图表
        @param chart_type: 图表类型目前支持只支持5种
        @param row_chart_num: 生成图表后插入的行
        @param col_chart_num: 生成图表后插入的列
        @param row_label_min_num: 图表图例的最小行
        @param row_label_max_num: 图表图例的最大行
        @param col_label_min_num: 图表图例的列号
        @param row_data_min_num: 图表数据最小行,从表头开始算
        @param row_data_max_num: 图表数据最大行
        @param col_data_min_num: 图表数据最小列
        @param col_data_max_num: 图表数据最大列,没有则不填
        @param chart_title: 图表标题,没有则不填
        @param chart_x_title: 图表x轴标题,没有则不填
        @param chart_y_title: 图表y轴标题,没有则不填
        @param chart_style_num: 图表样式数字,不填则用默认值
        @param input_sheet_name: 插入的sheet名,不填则使用当前使用的sheet
        @param color_list: 颜色列表,没有则不填,列表有值则需要与数据对应上
        @param symbol_str: 散点图使用,散点的样式,不填则默认
        @param smooth_bool: 是否光滑的线,默认不光滑
        @param chart_width_num: 图表长度
        @param chart_height_num: 图表高度
        @return: 无
        """

        def set_chart_line():
            # 下面两行为了让x轴内容多时显示正常,就写死
            chart_obj.y_axis.crossAx = 500
            chart_obj.x_axis = DateAxis(crossAx=100)
            data_obj = Reference(worksheet_obj, min_col=col_data_min_num,
                                 min_row=row_data_min_num, max_col=col_data_max_num, max_row=row_data_max_num)
            chart_obj.add_data(data_obj, titles_from_data=True)
            use_num = 0
            for line_obj in chart_obj.series:
                if color_list:
                    print(color_list[use_num])
                    line_obj.graphicalProperties.line.solidFill = color_list[use_num]
                # 是否光滑的线
                line_obj.smooth = smooth_bool
                use_num += 1
            chart_obj.set_categories(label_obj)

        def set_chart_bar():
            data_obj = Reference(worksheet_obj, min_col=col_data_min_num, min_row=row_data_min_num,
                                 max_col=col_data_max_num, max_row=row_data_max_num)
            chart_obj.add_data(data_obj, titles_from_data=True)
            use_num = 0
            for bar_obj in chart_obj.series:
                if color_list:
                    bar_obj.graphicalProperties.solidFill = color_list[use_num]
                use_num += 1
            chart_obj.set_categories(label_obj)

        def set_chart_doughnut():
            # 第几个值 向外突多少
            slices_list = list()
            for idx_num in range(row_data_max_num - row_data_min_num):
                data_point = DataPoint(idx=idx_num, explosion=2)
                if color_list:
                    data_point.graphicalProperties.solidFill = color_list[idx_num]
                slices_list.append(data_point)
            for col_data_use_num in range(col_data_min_num, col_data_max_num + 1):
                data_obj = Reference(worksheet_obj, min_col=col_data_use_num,
                                     min_row=row_data_min_num, max_row=row_data_max_num)
                series_obj = Series(data_obj, title_from_data=True)
                series_obj.data_points = slices_list
                chart_obj.series.append(series_obj)
            chart_obj.set_categories(label_obj)

        def set_chart_pie():
            # 第几个值 向外突多少
            slices_list = list()
            for idx_num in range(row_data_max_num - row_data_min_num):
                data_point = DataPoint(idx=idx_num, explosion=2)
                if color_list:
                    data_point.graphicalProperties.solidFill = color_list[idx_num]
                slices_list.append(data_point)

            data_obj = Reference(worksheet_obj, min_col=col_data_min_num,
                                 min_row=row_data_min_num, max_row=row_data_max_num)
            series_obj = Series(data_obj, title_from_data=True)
            series_obj.data_points = slices_list
            chart_obj.series.append(series_obj)
            chart_obj.set_categories(label_obj)

        def set_chart_scatter():
            use_num = 0
            for col_data_use_num in range(col_data_min_num, col_data_max_num + 1):
                data_obj = Reference(worksheet_obj, min_col=col_data_use_num,
                                     min_row=row_data_min_num, max_row=row_data_max_num)
                series_obj = Series(data_obj, label_obj, title_from_data=True)
                # 可以配置的类型有
                # 'star',方块 'dash',横线 'square',方块 'circle',圆 'picture', 'diamond',小菱形
                # 'dot',小横线 'plus',方块 'x',方块 'auto',菱形 'triangle'三角
                series_obj.marker.symbol = symbol_str
                series_obj.marker.graphicalProperties.solidFill = color_list[use_num]
                series_obj.marker.graphicalProperties.line.solidFill = color_list[use_num]
                # 不要有线
                series_obj.graphicalProperties.line.noFill = True
                chart_obj.series.append(series_obj)
                use_num += 1

        def add_common():
            # 标题头
            if chart_title is not None:
                chart_obj.title = chart_title
            # Y轴文字
            if chart_y_title is not None:
                chart_obj.y_axis.title = chart_y_title
            # X轴文字
            if chart_x_title is not None:
                chart_obj.x_axis.title = chart_x_title
            chart_obj.style = chart_style_num
            # 长度
            if chart_width_num:
                chart_obj.width = chart_width_num
            # 高度
            if chart_height_num:
                chart_obj.height = chart_height_num

        # 图表类型
        chart_dict = dict()
        # 线图
        chart_dict["line"] = [LineChart(), set_chart_line]
        # 柱状图
        chart_dict["bar"] = [BarChart(), set_chart_bar]
        # 甜甜圈图
        chart_dict["doughnut"] = [DoughnutChart(), set_chart_doughnut]
        # 饼图
        chart_dict["pie"] = [PieChart(), set_chart_pie]
        # 散点图
        chart_dict["scatter"] = [ScatterChart(), set_chart_scatter]

        if chart_type in chart_dict.keys():
            chart_obj = chart_dict[chart_type][0]
            worksheet_obj = self.get_sheet(input_sheet_name)
            label_obj = Reference(worksheet_obj, min_col=col_label_min_num,
                                  min_row=row_label_min_num, max_row=row_label_max_num)

            chart_dict.get(chart_type)[1]()
            add_common()
            # 图表插入的位置
            cell_chart_str = self.__get_cell_id(row_chart_num, col_chart_num)
            worksheet_obj.add_chart(chart_obj, cell_chart_str)
            self.result_str = "pass"
            self.reason_var = f"图表{chart_type},生成成功"
        else:
            self.result_str = "fail"
            self.reason_var = f"设置的图表类型不支持,设置的为:{chart_type}"

    def chart_line(self, row_chart_num, col_chart_num, row_label_min_num, row_label_max_num, col_label_min_num,
                   row_data_min_num, row_data_max_num, col_data_min_num, col_data_max_num,
                   chart_title=None, chart_x_title=None, chart_y_title=None, chart_style_num=10,
                   input_sheet_name=None, color_list=None, smooth_bool=False,
                   chart_width_num=None, chart_height_num=None):
        """
        @public: 生成线图
        @param row_chart_num: 生成图表后插入的行
        @param col_chart_num: 生成图表后插入的列
        @param row_label_min_num: 图表图例的最小行
        @param row_label_max_num: 图表图例的最大行
        @param col_label_min_num: 图表图例的列号
        @param row_data_min_num: 图表数据最小行,从表头开始算
        @param row_data_max_num: 图表数据最大行
        @param col_data_min_num: 图表数据最小列
        @param col_data_max_num: 图表数据最大列,没有则不填
        @param chart_title: 图表标题,没有则不填
        @param chart_x_title: 图表x轴标题,没有则不填
        @param chart_y_title: 图表y轴标题,没有则不填
        @param chart_style_num: 图表样式数字,不填则用默认值
        @param input_sheet_name: 插入的sheet名,不填则使用当前使用的sheet
        @param color_list: 颜色列表,没有则不填,列表有值则需要与数据对应上
        @param smooth_bool: 是否光滑的线,默认不光滑
        @param chart_width_num: 图表长度
        @param chart_height_num: 图表高度
        @return: 无
        """
        self.__make_chart("line", row_chart_num, col_chart_num,
                          row_label_min_num, row_label_max_num, col_label_min_num,
                          row_data_min_num, row_data_max_num, col_data_min_num, col_data_max_num,
                          chart_title, chart_x_title, chart_y_title, chart_style_num, input_sheet_name, color_list,
                          smooth_bool=smooth_bool, chart_width_num=chart_width_num, chart_height_num=chart_height_num)

    def chart_bar(self, row_chart_num, col_chart_num, row_label_min_num, row_label_max_num, col_label_min_num,
                  row_data_min_num, row_data_max_num, col_data_min_num, col_data_max_num,
                  chart_title=None, chart_x_title=None, chart_y_title=None, chart_style_num=10,
                  input_sheet_name=None, color_list=None, chart_width_num=None, chart_height_num=None):
        """
        @public: 生成柱状图
        @param row_chart_num: 生成图表后插入的行
        @param col_chart_num: 生成图表后插入的列
        @param row_label_min_num: 图表图例的最小行
        @param row_label_max_num: 图表图例的最大行
        @param col_label_min_num: 图表图例的列号
        @param row_data_min_num: 图表数据最小行,从表头开始算
        @param row_data_max_num: 图表数据最大行
        @param col_data_min_num: 图表数据最小列
        @param col_data_max_num: 图表数据最大列,没有则不填
        @param chart_title: 图表标题,没有则不填
        @param chart_x_title: 图表x轴标题,没有则不填
        @param chart_y_title: 图表y轴标题,没有则不填
        @param chart_style_num: 图表样式数字,不填则用默认值
        @param input_sheet_name: 插入的sheet名,不填则使用当前使用的sheet
        @param color_list: 颜色列表,没有则不填,列表有值则需要与数据对应上
        @param chart_width_num: 图表长度
        @param chart_height_num: 图表高度
        @return: 无
        """
        self.__make_chart("bar", row_chart_num, col_chart_num,
                          row_label_min_num, row_label_max_num, col_label_min_num,
                          row_data_min_num, row_data_max_num, col_data_min_num, col_data_max_num,
                          chart_title, chart_x_title, chart_y_title, chart_style_num, input_sheet_name,
                          color_list, chart_width_num=chart_width_num, chart_height_num=chart_height_num)

    def chart_doughnut(self, row_chart_num, col_chart_num, row_label_min_num, row_label_max_num, col_label_min_num,
                       row_data_min_num, row_data_max_num, col_data_min_num, col_data_max_num,
                       chart_title=None, chart_x_title=None, chart_y_title=None, chart_style_num=10,
                       input_sheet_name=None, color_list=None, chart_width_num=None, chart_height_num=None):
        """
        @public: 生成甜甜圈图
        @param row_chart_num: 生成图表后插入的行
        @param col_chart_num: 生成图表后插入的列
        @param row_label_min_num: 图表图例的最小行
        @param row_label_max_num: 图表图例的最大行
        @param col_label_min_num: 图表图例的列号
        @param row_data_min_num: 图表数据最小行,从表头开始算
        @param row_data_max_num: 图表数据最大行
        @param col_data_min_num: 图表数据最小列
        @param col_data_max_num: 图表数据最大列,没有则不填
        @param chart_title: 图表标题,没有则不填
        @param chart_x_title: 图表x轴标题,没有则不填
        @param chart_y_title: 图表y轴标题,没有则不填
        @param chart_style_num: 图表样式数字,不填则用默认值
        @param input_sheet_name: 插入的sheet名,不填则使用当前使用的sheet
        @param color_list: 颜色列表,没有则不填,列表有值则需要与数据对应上
        @param chart_width_num: 图表长度
        @param chart_height_num: 图表高度
        @return:
        """
        self.__make_chart("doughnut", row_chart_num, col_chart_num,
                          row_label_min_num, row_label_max_num, col_label_min_num,
                          row_data_min_num, row_data_max_num, col_data_min_num, col_data_max_num,
                          chart_title, chart_x_title, chart_y_title, chart_style_num, input_sheet_name,
                          color_list, chart_width_num=chart_width_num, chart_height_num=chart_height_num)

    def chart_pie(self, row_chart_num, col_chart_num, row_label_min_num, row_label_max_num, col_label_min_num,
                  row_data_min_num, row_data_max_num, col_data_min_num,
                  chart_title=None, chart_x_title=None, chart_y_title=None, chart_style_num=10,
                  input_sheet_name=None, color_list=None, chart_width_num=None, chart_height_num=None):
        """
        @public: 生成饼图
        @param row_chart_num: 生成图表后插入的行
        @param col_chart_num: 生成图表后插入的列
        @param row_label_min_num: 图表图例的最小行
        @param row_label_max_num: 图表图例的最大行
        @param col_label_min_num: 图表图例的列号
        @param row_data_min_num: 图表数据最小行,从表头开始算
        @param row_data_max_num: 图表数据最大行
        @param col_data_min_num: 图表数据最小列
        @param chart_title: 图表标题,没有则不填
        @param chart_x_title: 图表x轴标题,没有则不填
        @param chart_y_title: 图表y轴标题,没有则不填
        @param chart_style_num: 图表样式数字,不填则用默认值
        @param input_sheet_name: 插入的sheet名,不填则使用当前使用的sheet
        @param color_list: 颜色列表,没有则不填,列表有值则需要与数据对应上
        @param chart_width_num: 图表长度
        @param chart_height_num: 图表高度
        @return:
        """
        self.__make_chart("pie", row_chart_num, col_chart_num,
                          row_label_min_num, row_label_max_num, col_label_min_num,
                          row_data_min_num, row_data_max_num, col_data_min_num, None,
                          chart_title, chart_x_title, chart_y_title, chart_style_num, input_sheet_name,
                          color_list, chart_width_num=chart_width_num, chart_height_num=chart_height_num)

    def chart_scatter(self, row_chart_num, col_chart_num, row_label_min_num, row_label_max_num, col_label_min_num,
                      row_data_min_num, row_data_max_num, col_data_min_num, col_data_max_num,
                      chart_title=None, chart_x_title=None, chart_y_title=None, chart_style_num=10,
                      input_sheet_name=None, color_list=None, symbol_str="circle",
                      chart_width_num=None, chart_height_num=None):
        """
        @public: 生成散点图
        @param row_chart_num: 生成图表后插入的行
        @param col_chart_num: 生成图表后插入的列
        @param row_label_min_num: 图表图例的最小行
        @param row_label_max_num: 图表图例的最大行
        @param col_label_min_num: 图表图例的列号
        @param row_data_min_num: 图表数据最小行,从表头开始算
        @param row_data_max_num: 图表数据最大行
        @param col_data_min_num: 图表数据最小列
        @param col_data_max_num: 图表数据最大列,没有则不填
        @param chart_title: 图表标题,没有则不填
        @param chart_x_title: 图表x轴标题,没有则不填
        @param chart_y_title: 图表y轴标题,没有则不填
        @param chart_style_num: 图表样式数字,不填则用默认值
        @param input_sheet_name: 插入的sheet名,不填则使用当前使用的sheet
        @param color_list: 颜色列表,没有则不填,列表有值则需要与数据对应上
        @param symbol_str: 散点的样式,不填则默认
        @param chart_width_num: 图表长度
        @param chart_height_num: 图表高度
        @return:
        """
        self.__make_chart("scatter", row_chart_num, col_chart_num,
                          row_label_min_num, row_label_max_num, col_label_min_num,
                          row_data_min_num, row_data_max_num, col_data_min_num, col_data_max_num,
                          chart_title, chart_x_title, chart_y_title, chart_style_num, input_sheet_name, color_list,
                          symbol_str, chart_width_num=chart_width_num, chart_height_num=chart_height_num)

    @staticmethod
    def get_color_code(cn_str=None):
        """
        @public: 读取颜色
        @param cn_str: 读取的key,如果不输入则返回所有
        @return: 单个颜色编码或全部颜色编码
        """
        color_dict = dict()
        # color_dict["白"] = "FFFFFF"
        # color_dict["黑"] = "000000"
        color_dict["红"] = "FF0000"
        color_dict["橙"] = "FFA500"
        color_dict["黄"] = "FFFF00"
        color_dict["绿"] = "008000"
        color_dict["青"] = "00FFFF"
        color_dict["蓝"] = "0000FF"
        color_dict["紫"] = "800080"
        color_dict["灰"] = "808080"
        # color_dict["银"] = "C0C0C0"
        color_dict["金"] = "FFD700"
        color_dict["粉"] = "FFC0CB"
        color_dict["棕"] = "A52A2A"
        color_dict["桃"] = "FFDAB9"
        color_dict["淡蓝"] = "ADD8E6"
        color_dict["淡粉"] = "FFB6C1"
        color_dict["淡青"] = "E1FFFF"
        color_dict["淡绿"] = "90EE90"
        color_dict["淡黄"] = "FFFFE0"
        color_dict["淡灰"] = "D3D3D3"
        if cn_str in color_dict.keys():
            en_var = color_dict[cn_str]
        else:
            en_var = color_dict.values()
        return en_var


if __name__ == "__main__":
    """
    rows_list = [
    ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
    [1, 40, 30, 25],
    [2, 40, 25, 30],
    [3, 50, 30, 45],
    [4, 30, 25, 40],
    [5, 25, 35, 30],
    [6, 20, 40, 35],
    ]
    excel_class = ExcelOpenPyXl()
    excel_class.modify_sheet("line")
    excel_class.write_list_cells(rows_list)
    title_str = excel_class.read_cell_value(1, 1)
    excel_class.chart_line(4, 8, 2, 7, 1, 1, 7, 2, 4, title_str, "uuuuuu", color_list=["FF0000", "FF0000", "FF0000"])
    excel_class.chart_scatter(8, 8, 2, 7, 1, 1, 7, 2, 4, title_str, "uuuuuu", color_list=["FF0000", "FF0000", "FF0000"])
    excel_class.save_excel("1.xlsx")
    excel_class.close_excel()
    """
